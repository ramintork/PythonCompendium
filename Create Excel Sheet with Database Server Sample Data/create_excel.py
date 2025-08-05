import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Color
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles.numbers import FORMAT_NUMBER_00

def create_excel_with_heatmaps(csv_file, excel_file):
    df = pd.read_csv(csv_file)

    # Create a new Excel workbook
    wb = Workbook()
    
    # Remove the default sheet
    wb.remove(wb.active)

    # Create Server Data sheet
    ws_data = wb.create_sheet("Server Data")
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws_data.append(row)

    # Format numeric columns to 2 decimal places
    all_numeric_cols = [col for col in df.columns if df[col].dtype in ["float64", "int64"]]

    for col_name in all_numeric_cols:
        col_idx = df.columns.get_loc(col_name) + 1  # +1 because Excel is 1-indexed
        col_letter = chr(64 + col_idx)  # Convert to letter
        for row in range(2, len(df) + 2):  # Skip header
            ws_data[f'{col_letter}{row}'].number_format = FORMAT_NUMBER_00

    # Function to create heatmap for a specific metric
    def create_metric_heatmap(metric_name):
        ws_heatmap = wb.create_sheet(f"{metric_name} Heatmap")
        
        # Create pivot table for the metric
        heatmap_data = df.pivot_table(values=metric_name, index='type', columns='region', aggfunc='mean')
        
        # Write the pivoted data to the sheet
        for r_idx, row in enumerate(dataframe_to_rows(heatmap_data, index=True, header=True), 1):
            ws_heatmap.append(row)

        # Format numbers to 2 decimal places
        for row in range(2, heatmap_data.shape[0] + 2):
            for col in range(2, heatmap_data.shape[1] + 2):
                col_letter = chr(64 + col)
                ws_heatmap[f'{col_letter}{row}'].number_format = FORMAT_NUMBER_00

        # Apply conditional formatting
        min_val = heatmap_data.min().min()
        max_val = heatmap_data.max().max()

        start_row = 2
        start_col = 2
        end_row = start_row + heatmap_data.shape[0] - 1
        end_col = start_col + heatmap_data.shape[1] - 1

        heatmap_range = f"B{start_row}:{chr(65 + end_col)}{end_row}"

        ws_heatmap.conditional_formatting.add(
            heatmap_range,
            ColorScaleRule(
                start_type='min',
                start_value=min_val,
                start_color=Color('FFEB84'),  # Light yellow
                mid_type='percentile',
                mid_value=50,
                mid_color=Color('FF8000'),  # Orange
                end_type='max',
                end_value=max_val,
                end_color=Color('990000')  # Dark red
            )
        )

        # Add legend
        legend_start_row = end_row + 3
        ws_heatmap[f'A{legend_start_row}'].value = f"Heatmap Legend ({metric_name} Utilization)"
        ws_heatmap[f'A{legend_start_row + 1}'].value = "Low"
        ws_heatmap[f'B{legend_start_row + 1}'].fill = PatternFill(start_color='FFEB84', end_color='FFEB84', fill_type='solid')
        ws_heatmap[f'A{legend_start_row + 2}'].value = "Medium"
        ws_heatmap[f'B{legend_start_row + 2}'].fill = PatternFill(start_color='FF8000', end_color='FF8000', fill_type='solid')
        ws_heatmap[f'A{legend_start_row + 3}'].value = "High"
        ws_heatmap[f'B{legend_start_row + 3}'].fill = PatternFill(start_color='990000', end_color='990000', fill_type='solid')

    # Create heatmaps for each metric
    metrics = ['CPU', 'IO', 'Storage', 'Memory']
    for metric in metrics:
        create_metric_heatmap(metric)

    # Create Overall Heatmap sheet
    ws_overall = wb.create_sheet("Overall Heatmap")
    
    # Write the entire DataFrame to the Overall Heatmap sheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws_overall.append(row)

    # Format numeric columns to 2 decimal places in Overall Heatmap
    for col_name in all_numeric_cols:
        col_idx = df.columns.get_loc(col_name) + 1
        col_letter = chr(64 + col_idx)
        for row in range(2, len(df) + 2):
            ws_overall[f'{col_letter}{row}'].number_format = FORMAT_NUMBER_00

    # Apply conditional formatting for each metric based on its threshold
    data_start_row = 2
    data_end_row = data_start_row + df.shape[0] - 1

    # Define the columns and their corresponding threshold columns
    metric_threshold_pairs = {
        'CPU': 'CPU_thresh',
        'IO': 'IO_thresh',
        'Storage': 'Storage_thresh',
        'Memory': 'Memory_thresh'
    }

    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    for metric, threshold_metric in metric_threshold_pairs.items():
        metric_col_idx = df.columns.get_loc(metric) + 1
        threshold_col_idx = df.columns.get_loc(threshold_metric) + 1
        
        metric_col_letter = chr(64 + metric_col_idx)
        threshold_col_letter = chr(64 + threshold_col_idx)
        
        # Apply conditional formatting using FormulaRule
        # This formula checks if the current cell's value (e.g., CPU) is greater than its corresponding threshold (e.g., CPU_thresh)
        # The $ before the column letter makes the column reference absolute, while the row reference is relative.
        # This ensures the formula correctly compares each metric value with its corresponding threshold in the same row.
        formula = f'${metric_col_letter}2 > ${threshold_col_letter}2'
        
        # Define the range for conditional formatting. It should cover all data rows for the specific metric column.
        cf_range = f'{metric_col_letter}{data_start_row}:{metric_col_letter}{data_end_row}'
        
        ws_overall.conditional_formatting.add(
            cf_range,
            FormulaRule(
                formula=[formula],
                fill=red_fill
            )
        )

    # Add legend for Overall Heatmap
    legend_start_row = data_end_row + 3
    ws_overall[f'A{legend_start_row}'].value = "Overall Heatmap Legend"
    ws_overall[f'A{legend_start_row + 1}'].value = "Red cells indicate metrics exceeding their thresholds"
    ws_overall[f'B{legend_start_row + 1}'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Save the workbook
    wb.save(excel_file)
    print(f"Excel file '{excel_file}' created successfully with multiple heatmaps and formatting.")

if __name__ == "__main__":
    create_excel_with_heatmaps("server_data.csv", "server_performance.xlsx")


